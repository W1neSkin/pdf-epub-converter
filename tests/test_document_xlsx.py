import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from document_content import build_page_records  # noqa: E402
from document_xlsx import write_document_xlsx  # noqa: E402
import table_extractor  # noqa: E402


class FakePage:
    """Return positioned words like pdfplumber does."""

    @staticmethod
    def extract_words(**_kwargs):
        return [
            {"text": "Кредитный", "x0": 10, "x1": 70, "top": 10, "bottom": 20},
            {"text": "отчёт", "x0": 75, "x1": 110, "top": 10, "bottom": 20},
            {"text": "Имя", "x0": 10, "x1": 35, "top": 50, "bottom": 60},
            {"text": "Сумма", "x0": 110, "x1": 150, "top": 50, "bottom": 60},
            {"text": "Конец", "x0": 10, "x1": 45, "top": 120, "bottom": 130},
        ]


def sample_records():
    """Create one page containing narrative text and a table."""
    return build_page_records(
        FakePage(),
        3,
        [
            {
                "table_index": 1,
                "strategy": "lines_strict",
                "bbox": [0, 40, 200, 100],
                "rows": [["Имя", "Сумма"], ["Иван", "100"]],
            }
        ],
    )


def test_page_records_keep_order_without_repeating_table_words():
    records = sample_records()

    assert [record["content_type"] for record in records] == [
        "text",
        "table",
        "table",
        "text",
    ]
    narrative = " ".join(
        str(record["text"])
        for record in records
        if record["content_type"] == "text"
    )
    assert narrative == "Кредитный отчёт Конец"
    assert "Имя" not in narrative


def test_workbook_is_readable_and_contains_no_technical_columns(tmp_path):
    output_path = tmp_path / "document.xlsx"

    assert write_document_xlsx(sample_records(), str(output_path), page_count=3) == 4

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Page 1", "Page 2", "Page 3"]
    sheet = workbook["Page 3"]
    values = [
        str(cell.value)
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "Кредитный отчёт" in values
    assert "Имя" in values
    assert "Сумма" in values
    assert "content_type" not in values


def test_export_job_creates_workbook_and_both_table_downloads(monkeypatch, tmp_path):
    """One extraction should expose a workbook and both table formats."""
    pdf_path = tmp_path / "source.pdf"
    pdf_path.write_bytes(b"test")
    output_dir = tmp_path / "output"
    result = {
        "pages": 1,
        "tables": [
            {
                "page": 1,
                "table_index": 1,
                "strategy": "lines_strict",
                "rows": [["Имя", "Сумма"], ["Иван", "100"]],
            }
        ],
        "document_records": [
            {
                "page": 1,
                "content_order": 1,
                "content_type": "text",
                "text": "Кредитный отчёт",
                "columns": [],
            },
            {
                "page": 1,
                "content_order": 2,
                "content_type": "table",
                "table_index": 1,
                "row_index": 1,
                "strategy": "lines_strict",
                "text": "",
                "columns": ["Имя", "Сумма"],
            },
        ],
        "strategy_summary": {"lines_strict": 1},
        "used_camelot": False,
        "used_ocr": False,
    }
    monkeypatch.setattr(table_extractor, "extract_tables_from_pdf", lambda _path: result)
    monkeypatch.setenv("PUBLIC_API_URL", "https://converter.example")

    table_extractor.run_table_extraction_job(
        "job-1",
        str(pdf_path),
        str(output_dir),
        "Отчёт.pdf",
    )

    status = json.loads((output_dir / "status.json").read_text(encoding="utf-8"))
    assert status["status"] == "completed"
    assert status["document_row_count"] == 2
    assert status["download_name"] == "Отчёт_document.xlsx"
    assert status["output_kind"] == "xlsx"
    assert status["tables_download_url"].endswith("?kind=tables")
    assert status["archive_download_url"].endswith("?kind=archive")
    assert (output_dir / status["output_filename"]).exists()
    assert (output_dir / status["tables_filename"]).exists()
    assert (output_dir / status["archive_filename"]).exists()


def test_export_job_succeeds_when_pdf_has_text_but_no_tables(monkeypatch, tmp_path):
    """A full-document workbook must not require tables."""
    pdf_path = tmp_path / "source.pdf"
    pdf_path.write_bytes(b"test")
    output_dir = tmp_path / "output"
    result = {
        "pages": 1,
        "tables": [],
        "document_records": [
            {
                "page": 1,
                "content_order": 1,
                "content_type": "text",
                "text": "Plain document text",
                "columns": [],
            }
        ],
        "strategy_summary": {},
        "used_camelot": False,
        "used_ocr": False,
    }
    monkeypatch.setattr(table_extractor, "extract_tables_from_pdf", lambda _path: result)

    table_extractor.run_table_extraction_job(
        "job-2",
        str(pdf_path),
        str(output_dir),
        "plain.pdf",
    )

    status = json.loads((output_dir / "status.json").read_text(encoding="utf-8"))
    assert status["status"] == "completed"
    assert status["table_count"] == 0
    assert status["tables_download_url"] is None
    assert status["archive_download_url"] is None
