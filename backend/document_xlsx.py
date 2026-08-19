"""Create a human-readable Excel workbook from ordered PDF content."""

import os
from collections import defaultdict
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TABLE_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)
HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")


def _clean(value: object) -> str:
    """Remove control characters that Excel cannot store."""
    return ILLEGAL_CHARACTERS_RE.sub("", str(value or "")).strip()


def _set_column_widths(sheet, widths: Dict[int, int], minimum_columns: int) -> None:
    """Apply practical widths so content is readable immediately."""
    for column_index in range(1, minimum_columns + 1):
        width = widths.get(column_index, 14)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            45,
            max(14, width + 2),
        )


def _write_page(sheet, records: List[Dict[str, object]]) -> None:
    """Render one PDF page as readable text and bordered table rows."""
    max_table_columns = max(
        (len(record.get("columns") or []) for record in records),
        default=1,
    )
    content_columns = max(6, max_table_columns)
    widths: Dict[int, int] = defaultdict(int)
    next_row = 1
    previous_group = None

    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    for record in records:
        content_type = str(record.get("content_type") or "")
        table_index = int(record.get("table_index", 0) or 0)
        current_group = (
            ("table", table_index)
            if content_type == "table"
            else ("text", 0)
        )
        # Keep the workbook compact. One separator row is enough to show
        # where narrative text ends or a different table begins.
        if previous_group is not None and current_group != previous_group:
            next_row += 1
        row_number = next_row

        if record.get("content_type") == "text":
            text = _clean(record.get("text"))
            if text:
                sheet.merge_cells(
                    start_row=row_number,
                    start_column=1,
                    end_row=row_number,
                    end_column=content_columns,
                )
                cell = sheet.cell(row=row_number, column=1, value=text)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(
                    name="Aptos",
                    size=14 if row_number <= 3 else 11,
                    bold=row_number <= 3,
                )
                sheet.row_dimensions[row_number].height = 24
                widths[1] = max(widths[1], min(len(text), 45))
        else:
            columns = list(record.get("columns") or [])
            is_header = int(record.get("row_index", 0) or 0) == 1
            for column_index, value in enumerate(columns, 1):
                text = _clean(value)
                cell = sheet.cell(
                    row=row_number,
                    column=column_index,
                    value=text,
                )
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = TABLE_BORDER
                cell.font = Font(name="Aptos", size=10, bold=is_header)
                if is_header:
                    cell.fill = HEADER_FILL
                widths[column_index] = max(
                    widths[column_index],
                    min(max((len(line) for line in text.splitlines()), default=0), 43),
                )
            sheet.row_dimensions[row_number].height = 21
        next_row = row_number + 1
        previous_group = current_group

    _set_column_widths(sheet, widths, content_columns)


def write_document_xlsx(
    records: List[Dict[str, object]],
    output_path: str,
    page_count: int = 0,
) -> int:
    """Write one readable worksheet per PDF page and return record count."""
    if not records:
        raise ValueError("No text or tables found in this PDF")

    pages: Dict[int, List[Dict[str, object]]] = defaultdict(list)
    for record in records:
        page_number = int(record.get("page", 0) or 0)
        pages[page_number].append(record)

    workbook = Workbook()
    workbook.remove(workbook.active)
    total_pages = max(page_count, max(pages, default=1))
    for page_number in range(1, total_pages + 1):
        sheet = workbook.create_sheet(title=f"Page {page_number}")
        _write_page(sheet, pages.get(page_number, []))

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    workbook.save(output_path)
    return len(records)
